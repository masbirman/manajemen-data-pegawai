from fastapi import APIRouter
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/template", tags=["template"])


@router.get("/download")
async def download_template():
    """
    Download template Excel file for employee data upload.
    
    Returns:
        Excel (.xlsx) file with required columns and example data
    """
    # Create a new workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pegawai"
    
    # Define header style
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header
    headers = [
        'NIP',
        'Nama',
        'NIK',
        'NPWP',
        'Tanggal Lahir',
        'Kode Bank',
        'Nama Bank',
        'Nomor Rekening'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write example data
    example_data = [
        ['EMP001', 'Budi Santoso', '3201012345678901', '123456789012345', '1990-05-15', 'BCA', 'BCA', '1234567890'],
        ['EMP002', 'Siti Nurhaliza', '3201023456789012', '234567890123456', '1985-08-20', 'MDR', 'Mandiri', '2345678901'],
        ['EMP003', 'Ahmad Hidayat', '3201034567890123', '345678901234567', '1992-03-10', 'BRI', 'BRI', '3456789012'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    column_widths = [15, 25, 20, 20, 18, 12, 15, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create streaming response
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename=template_data_pegawai.xlsx'
        }
    )
